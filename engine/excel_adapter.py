import io
import logging

from .base_adapter import BaseFormatAdapter
from .adapter_registry import register_adapter

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _get_sheet(wb, sheet_name):
    """Return the requested worksheet or the active one."""
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.active


def _cell_value(cell):
    """Extract a usable value from a cell, handling merged cells."""
    if cell is None:
        return ''
    value = cell.value
    if value is None:
        return ''
    return value


@register_adapter
class ExcelAdapter(BaseFormatAdapter):

    FORMAT_KEY = 'xlsx'
    DISPLAY_NAME = 'Excel (.xlsx)'
    FILE_EXTENSIONS = ['.xlsx', '.xls']
    MIME_TYPES = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
    ]
    PYTHON_DEPENDENCIES = ['openpyxl']

    def parse(self, file_data, options):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is required to parse Excel files")

        sheet_name = options.get('sheet_name')
        header_row = int(options.get('header_row', 1))
        data_start_row = int(options.get('data_start_row', header_row + 1))

        wb = openpyxl.load_workbook(
            io.BytesIO(file_data), read_only=True, data_only=True,
        )
        try:
            ws = _get_sheet(wb, sheet_name)
            if ws is None:
                return

            headers = []
            for row_idx, row in enumerate(ws.iter_rows(), 1):
                if row_idx == header_row:
                    for cell in row:
                        val = _cell_value(cell)
                        headers.append(str(val).strip() if val != '' else f'col_{len(headers)}')
                    continue
                if row_idx < data_start_row:
                    continue

                # Skip completely empty rows
                values = [_cell_value(c) for c in row]
                if not any(v != '' for v in values):
                    continue

                row_dict = {}
                for i, value in enumerate(values):
                    key = headers[i] if i < len(headers) else f'col_{i}'
                    row_dict[key] = value
                yield row_dict
        finally:
            wb.close()

    def write(self, records, fields, options):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is required to write Excel files")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = options.get('sheet_name', 'Sheet1')

        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Write headers
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Track max widths for auto-sizing (initialise from header lengths)
        col_widths = [len(str(f)) for f in fields]

        # Write data rows
        row_num = 2
        for record in records:
            for col_idx, field in enumerate(fields, 1):
                value = record.get(field, '')
                ws.cell(row=row_num, column=col_idx, value=value)
                val_len = len(str(value)) if value is not None else 0
                if col_idx - 1 < len(col_widths):
                    col_widths[col_idx - 1] = max(col_widths[col_idx - 1], val_len)
            row_num += 1

        # Apply auto-width (add padding)
        for col_idx, width in enumerate(col_widths, 1):
            adjusted = min(width + 4, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        output = io.BytesIO()
        wb.save(output)
        wb.close()
        return output.getvalue()

    def detect_columns(self, file_data, options):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl is required to read Excel files")

        sheet_name = options.get('sheet_name')
        header_row = int(options.get('header_row', 1))

        wb = openpyxl.load_workbook(
            io.BytesIO(file_data), read_only=True, data_only=True,
        )
        try:
            ws = _get_sheet(wb, sheet_name)
            if ws is None:
                return []

            for row_idx, row in enumerate(ws.iter_rows(), 1):
                if row_idx == header_row:
                    headers = []
                    for cell in row:
                        val = _cell_value(cell)
                        headers.append(str(val).strip() if val != '' else f'col_{len(headers)}')
                    return headers
            return []
        finally:
            wb.close()
